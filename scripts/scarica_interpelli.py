#!/usr/bin/env python3
"""
Scraper e parser degli Interpelli dell'Agenzia delle Entrate (2024-2025).
Scarica i PDF dal sito AdE, estrae le sezioni strutturate e genera un database JSON.

Uso:
    python scarica_interpelli.py

Il file Excel "INTERPELLI IVA.xlsx" deve trovarsi nella cartella padre (../INTERPELLI IVA.xlsx).
Output: interpelli_2024_2025_database.json nella stessa cartella padre.
"""

import os
import re
import json
import time
import sys
from pathlib import Path
from datetime import datetime

import requests
import pdfplumber
import openpyxl

# ─── Configurazione ─────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
BASE_DIR = SCRIPT_DIR.parent  # cartella FISCO
EXCEL_PATH = BASE_DIR / "INTERPELLI IVA.xlsx"
OUTPUT_JSON = BASE_DIR / "interpelli_2024_2025_database.json"
PDF_CACHE_DIR = SCRIPT_DIR / "pdf_cache"
ERRORS_LOG = SCRIPT_DIR / "errori_download.json"

# Pausa tra download (secondi) — per non sovraccaricare il server AdE
DOWNLOAD_DELAY = 1.0
# Timeout download (secondi)
DOWNLOAD_TIMEOUT = 30
# Riprova download N volte
MAX_RETRIES = 3

# Headers per simulare un browser reale
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/pdf,*/*',
    'Accept-Language': 'it-IT,it;q=0.9,en-US;q=0.8,en;q=0.7',
}


# ─── Step 1: Estrai metadati dall'Excel ─────────────────────────────────────

def extract_excel_metadata():
    """Estrae i metadati degli interpelli 2024-2025 dall'Excel."""
    print(f"📂 Leggo Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb.active

    records = []
    for row_idx in range(2, ws.max_row + 1):
        numero_cell = ws.cell(row=row_idx, column=2)
        data_cell = ws.cell(row=row_idx, column=3)
        tag_cell = ws.cell(row=row_idx, column=4)
        oggetto_cell = ws.cell(row=row_idx, column=5)
        massima_cell = ws.cell(row=row_idx, column=6)

        val = str(numero_cell.value or '').strip()
        if not val or val.startswith('="') or val in ('N°', 'None', ''):
            continue
        try:
            num = int(float(val))
        except (ValueError, TypeError):
            continue

        data = data_cell.value
        if not isinstance(data, datetime):
            continue
        if data.year not in (2024, 2025):
            continue

        link = numero_cell.hyperlink.target if numero_cell.hyperlink else None

        records.append({
            'numero': num,
            'anno': data.year,
            'data': data.strftime('%Y-%m-%d'),
            'tag': str(tag_cell.value or '').strip(),
            'oggetto': str(oggetto_cell.value or '').strip(),
            'massima': str(massima_cell.value or '').strip(),
            'link_pdf': link,
        })

    wb.close()
    print(f"   Trovati {len(records)} interpelli (2024: {sum(1 for r in records if r['anno']==2024)}, 2025: {sum(1 for r in records if r['anno']==2025)})")
    return records


# ─── Step 2: Scarica i PDF ──────────────────────────────────────────────────

def download_pdf(url, dest_path):
    """Scarica un singolo PDF con retry."""
    if dest_path.exists() and dest_path.stat().st_size > 1000:
        return True  # già scaricato

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=DOWNLOAD_TIMEOUT, stream=True)
            if resp.status_code == 200:
                with open(dest_path, 'wb') as f:
                    for chunk in resp.iter_content(8192):
                        f.write(chunk)
                if dest_path.stat().st_size > 500:
                    return True
                else:
                    dest_path.unlink()
            else:
                if attempt == MAX_RETRIES:
                    return False
        except (requests.RequestException, OSError) as e:
            if attempt == MAX_RETRIES:
                return False
        time.sleep(DOWNLOAD_DELAY)

    return False


def download_all_pdfs(records):
    """Scarica tutti i PDF degli interpelli."""
    PDF_CACHE_DIR.mkdir(exist_ok=True)

    total = len(records)
    downloaded = 0
    skipped = 0
    failed = []

    print(f"\n📥 Download di {total} PDF...")
    print(f"   Cache: {PDF_CACHE_DIR}")
    print(f"   (i PDF già scaricati verranno saltati)\n")

    for i, rec in enumerate(records):
        pdf_name = f"interpello_{rec['anno']}_{rec['numero']:04d}.pdf"
        pdf_path = PDF_CACHE_DIR / pdf_name
        rec['_pdf_path'] = str(pdf_path)

        if pdf_path.exists() and pdf_path.stat().st_size > 1000:
            skipped += 1
            downloaded += 1
            progress = (i + 1) / total * 100
            print(f"\r   [{progress:5.1f}%] {i+1}/{total} — ✓ {pdf_name} (cache)", end='', flush=True)
            continue

        if not rec['link_pdf']:
            failed.append({'numero': rec['numero'], 'anno': rec['anno'], 'errore': 'link mancante'})
            print(f"\r   [{(i+1)/total*100:5.1f}%] {i+1}/{total} — ✗ {pdf_name} (no link)", end='', flush=True)
            continue

        ok = download_pdf(rec['link_pdf'], pdf_path)
        if ok:
            downloaded += 1
            print(f"\r   [{(i+1)/total*100:5.1f}%] {i+1}/{total} — ✓ {pdf_name}", end='', flush=True)
        else:
            failed.append({'numero': rec['numero'], 'anno': rec['anno'], 'errore': f'HTTP error', 'url': rec['link_pdf']})
            print(f"\r   [{(i+1)/total*100:5.1f}%] {i+1}/{total} — ✗ {pdf_name} (ERRORE)", end='', flush=True)

        time.sleep(DOWNLOAD_DELAY)

    print(f"\n\n   Scaricati: {downloaded}/{total} | Dalla cache: {skipped} | Errori: {len(failed)}")

    if failed:
        with open(ERRORS_LOG, 'w') as f:
            json.dump(failed, f, ensure_ascii=False, indent=2)
        print(f"   Errori salvati in: {ERRORS_LOG}")

    return records, failed


# ─── Step 3: Parsa i PDF ────────────────────────────────────────────────────

def extract_pdf_text(pdf_path):
    """Estrae il testo completo da un PDF."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages = []
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    pages.append(t)
            return '\n'.join(pages)
    except Exception:
        return None


def parse_interpello_sections(text):
    """
    Parsa le sezioni tipiche di un interpello AdE:
    - OGGETTO
    - QUESITO
    - SOLUZIONE INTERPRETATIVA PROSPETTATA DAL CONTRIBUENTE
    - PARERE DELL'AGENZIA DELLE ENTRATE
    """
    if not text:
        return {}

    sections = {}

    # Pulizia testo
    text = re.sub(r'Divisione .+?\n', '', text)
    text = re.sub(r'Direzione .+?\n', '', text)
    text = re.sub(r'_{3,}', '', text)
    text = re.sub(r'-{3,}', '', text)

    # Pattern per identificare le sezioni — con varianti comuni
    section_patterns = [
        (r'(?:OGGETTO|Oggetto)\s*[:\.]?\s*', 'oggetto'),
        (r'(?:QUESITO|Quesito)\s*[:\.]?\s*', 'quesito'),
        (r'(?:SOLUZIONE INTERPRETATIVA PROSPETTATA DAL (?:CONTRIBUENTE|ISTANTE)|Soluzione interpretativa prospettata dal (?:contribuente|istante))\s*[:\.]?\s*', 'soluzione_contribuente'),
        (r'(?:PARERE DELL.AGENZIA DELLE ENTRATE|Parere dell.Agenzia delle [Ee]ntrate|PARERE DELL.AGENZIA)\s*[:\.]?\s*', 'parere_ade'),
    ]

    # Trova le posizioni di tutte le sezioni
    found_positions = []
    for pattern, key in section_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            found_positions.append((match.start(), match.end(), key))

    # Ordina per posizione
    found_positions.sort(key=lambda x: x[0])

    # Estrai il testo di ogni sezione
    for i, (start, content_start, key) in enumerate(found_positions):
        if i + 1 < len(found_positions):
            end = found_positions[i + 1][0]
        else:
            end = len(text)
        section_text = text[content_start:end].strip()
        # Pulizia
        section_text = re.sub(r'\n{3,}', '\n\n', section_text)
        sections[key] = section_text

    return sections


def extract_normative_references(text):
    """Estrae i riferimenti normativi citati nell'interpello."""
    if not text:
        return []

    refs = set()

    # DPR 633/72 (vecchio IVA)
    if re.search(r'd\.?P\.?R\.?\s*(?:n\.\s*)?633|decreto.+?(?:26\s+ottobre\s+1972|n\.\s*633)', text, re.IGNORECASE):
        arts = re.findall(r'(?:articol[oi]|art\.)\s*(\d+(?:\s*-\s*(?:bis|ter|quater|quinquies|sexies|septies|octies|novies|decies))?)\s*(?:del|,)?\s*(?:d\.?P\.?R\.?\s*(?:n\.\s*)?633|decreto.+?633)', text, re.IGNORECASE)
        for a in arts:
            refs.add(f"DPR 633/1972 art. {re.sub(r's+', '', a)}")

    # DL 331/93
    if re.search(r'd\.?[Ll]\.?\s*(?:n\.\s*)?331|decreto.+?(?:30\s+agosto\s+1993|n\.\s*331)', text, re.IGNORECASE):
        arts = re.findall(r'(?:articol[oi]|art\.)\s*(\d+(?:\s*-\s*(?:bis|ter|quater|quinquies|sexies|septies|octies|novies|decies))?)\s*(?:del|,)?\s*(?:d\.?[Ll]\.?\s*(?:n\.\s*)?331|decreto.+?331)', text, re.IGNORECASE)
        for a in arts:
            refs.add(f"DL 331/1993 art. {re.sub(r's+', '', a)}")

    # D.Lgs 10/2026 (nuovo TU IVA) - se qualche interpello lo cita già
    if re.search(r'd\.?\s*[Ll]gs\.?\s*(?:n\.\s*)?10.+?2026|testo\s+unico\s+iva', text, re.IGNORECASE):
        refs.add("D.Lgs. 10/2026 (TU IVA)")

    # Pattern generico per articoli citati senza norma specifica
    generic_arts = re.findall(
        r'(?:articol[oi]|art\.)\s*(\d+(?:\s*-\s*(?:bis|ter|quater|quinquies|sexies|septies|octies|novies|decies))?)',
        text, re.IGNORECASE
    )

    return {
        'riferimenti_specifici': sorted(refs),
        'articoli_citati': sorted(set(re.sub(r'\s+', '', a) for a in generic_arts))
    }


def extract_iva_topics(text, tag, oggetto):
    """Classifica l'interpello per temi IVA (allineato al TU IVA database)."""
    combined = f"{tag} {oggetto} {text or ''}".lower()
    topics = []

    topic_patterns = {
        'aliquote': [r'aliquot[ae]', r'\d+\s*per\s*cento', r'\d+%'],
        'esenzioni': [r'esent[ei]', r'esenzione', r'art\.?\s*10\b'],
        'detrazione': [r'detrazion[ei]', r'detraibil[ei]', r'pro[\s-]*rata'],
        'base_imponibile': [r'base\s+imponibile', r'corrispettivo'],
        'fatturazione': [r'fattur[ae]', r'fatturazione', r'nota\s+di\s+variazione'],
        'registrazione': [r'registr[oi]', r'registrazione'],
        'dichiarazione': [r'dichiarazion[ei]'],
        'rimborsi': [r'rimbors[oi]', r'credito\s+iva'],
        'operazioni_intra': [r'intra(?:unional|comunitari)', r'acquist[oi]\s+intra'],
        'importazioni': [r'importazion[ei]', r'dogan'],
        'esportazioni': [r'esportazion[ei]', r'plafond'],
        'regime_speciale': [r'regime\s+special', r'regime\s+forfet', r'margine', r'regime\s+agric'],
        'cessioni_beni': [r'cession[ei]\s+di\s+beni', r'cessione\s+immobil'],
        'prestazioni_servizi': [r'prestazion[ei]\s+di\s+servizi'],
        'reverse_charge': [r'reverse\s+charge', r'inversione\s+contabile'],
        'split_payment': [r'split\s+payment', r'scissione\s+dei\s+pagamenti'],
        'iva_edilizia': [r'edilizi', r'costruzion', r'ristrutturazion', r'superbonus', r'bonus\s+faccat'],
        'commercio_elettronico': [r'commercio\s+elettronico', r'e[\s-]*commerce', r'piattaforma'],
        'gruppo_iva': [r'gruppo\s+iva'],
        'territorialita': [r'territorialit', r'stabile\s+organizzazione'],
        'compensazioni': [r'compensazion'],
        'cessione_credito': [r'cessione\s+del?\s+credito', r'sconto\s+in\s+fattura'],
    }

    for topic, patterns in topic_patterns.items():
        for p in patterns:
            if re.search(p, combined):
                topics.append(topic)
                break

    return topics


# ─── Step 4: Costruisci il database ─────────────────────────────────────────

def build_database(records):
    """Costruisce il database JSON completo."""
    interpelli_db = []
    parse_errors = []

    total = len(records)
    print(f"\n📄 Parsing di {total} PDF...\n")

    for i, rec in enumerate(records):
        pdf_path = rec.get('_pdf_path')
        progress = (i + 1) / total * 100
        print(f"\r   [{progress:5.1f}%] {i+1}/{total} — Interpello {rec['numero']}/{rec['anno']}", end='', flush=True)

        # Estrai testo dal PDF
        full_text = None
        sections = {}
        norm_refs = {'riferimenti_specifici': [], 'articoli_citati': []}

        if pdf_path and Path(pdf_path).exists():
            full_text = extract_pdf_text(pdf_path)
            if full_text:
                sections = parse_interpello_sections(full_text)
                norm_refs = extract_normative_references(full_text)
            else:
                parse_errors.append({'numero': rec['numero'], 'anno': rec['anno'], 'errore': 'testo non estratto'})

        # Classifica temi
        topics = extract_iva_topics(
            full_text or rec['massima'],
            rec['tag'],
            rec['oggetto']
        )

        # Costruisci entry
        entry = {
            'id': f"interpello_{rec['anno']}_{rec['numero']}",
            'numero': rec['numero'],
            'anno': rec['anno'],
            'data': rec['data'],
            'tag': rec['tag'],
            'oggetto': rec['oggetto'],
            'massima': rec['massima'],
            'link_pdf': rec['link_pdf'],

            # Sezioni estratte dal PDF
            'sezioni': {
                'oggetto_completo': sections.get('oggetto', None),
                'quesito': sections.get('quesito', None),
                'soluzione_contribuente': sections.get('soluzione_contribuente', None),
                'parere_ade': sections.get('parere_ade', None),
            },

            # Testo integrale per ricerca full-text
            'testo_integrale': full_text,

            # Riferimenti normativi
            'riferimenti_normativi': norm_refs,

            # Classificazione tematica (allineata al TU IVA)
            'temi': topics,

            # Metadati per RAG
            'metadati_rag': {
                'search_text': _build_search_text(rec, sections),
                'citazione': f"Risposta a interpello n. {rec['numero']}/{rec['anno']} del {rec['data']}",
                'citazione_breve': f"Interpello {rec['numero']}/{rec['anno']}",
                'ha_testo_completo': full_text is not None,
                'lunghezza_caratteri': len(full_text) if full_text else 0,
            }
        }

        interpelli_db.append(entry)

    print(f"\n\n   Parsati: {total} | Errori parsing: {len(parse_errors)}")

    # Costruisci indice tematico
    indice_tematico = {}
    for entry in interpelli_db:
        for topic in entry['temi']:
            if topic not in indice_tematico:
                indice_tematico[topic] = []
            indice_tematico[topic].append({
                'id': entry['id'],
                'numero': entry['numero'],
                'anno': entry['anno'],
                'oggetto': entry['oggetto'],
            })

    # Database completo
    database = {
        'metadata': {
            'descrizione': 'Database interpelli Agenzia delle Entrate 2024-2025 per sistema RAG',
            'data_generazione': datetime.now().isoformat(),
            'totale_interpelli': len(interpelli_db),
            'per_anno': {
                '2024': sum(1 for e in interpelli_db if e['anno'] == 2024),
                '2025': sum(1 for e in interpelli_db if e['anno'] == 2025),
            },
            'per_tag': _count_by(interpelli_db, 'tag'),
            'con_testo_completo': sum(1 for e in interpelli_db if e['metadati_rag']['ha_testo_completo']),
            'note_per_rag': {
                'collegamento_tu_iva': 'Usare il campo riferimenti_normativi.riferimenti_specifici per collegare gli interpelli agli articoli del TU IVA tramite la mappatura_vecchio_nuovo_codice del database TU IVA',
                'retrieval_strategy': [
                    '1. Ricerca semantica su metadati_rag.search_text per trovare interpelli pertinenti alla domanda',
                    '2. Filtraggio per temi (allineati a quelli del TU IVA) per precision',
                    '3. Cross-reference con il TU IVA tramite riferimenti_normativi',
                    '4. Reranking basato su data (interpelli più recenti = più rilevanti)',
                ],
            },
        },
        'interpelli': interpelli_db,
        'indice_tematico': indice_tematico,
    }

    return database


def _build_search_text(rec, sections):
    """Costruisce il testo ottimizzato per embedding/ricerca semantica."""
    parts = [
        rec['oggetto'],
        rec['massima'],
    ]
    if sections.get('quesito'):
        # Prime 500 chars del quesito
        parts.append(sections['quesito'][:500])
    if sections.get('parere_ade'):
        # Prime 500 chars del parere
        parts.append(sections['parere_ade'][:500])

    return ' '.join(p for p in parts if p)


def _count_by(entries, field):
    """Conta occorrenze per campo."""
    counts = {}
    for e in entries:
        val = e[field]
        counts[val] = counts.get(val, 0) + 1
    return dict(sorted(counts.items(), key=lambda x: -x[1]))


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("SCRAPER INTERPELLI AGENZIA DELLE ENTRATE")
    print("=" * 60)

    if not EXCEL_PATH.exists():
        print(f"\n❌ File Excel non trovato: {EXCEL_PATH}")
        print(f"   Assicurati che 'INTERPELLI IVA.xlsx' sia nella cartella: {BASE_DIR}")
        sys.exit(1)

    # Step 1: Leggi Excel
    records = extract_excel_metadata()

    # Step 2: Scarica PDF
    records, download_errors = download_all_pdfs(records)

    # Step 3: Parsa e costruisci database
    database = build_database(records)

    # Step 4: Salva
    print(f"\n💾 Salvataggio in: {OUTPUT_JSON}")
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(database, f, ensure_ascii=False, indent=2)

    size_mb = OUTPUT_JSON.stat().st_size / 1024 / 1024

    print(f"\n{'=' * 60}")
    print(f"✅ COMPLETATO!")
    print(f"{'=' * 60}")
    print(f"   File: {OUTPUT_JSON}")
    print(f"   Dimensione: {size_mb:.1f} MB")
    print(f"   Interpelli: {database['metadata']['totale_interpelli']}")
    print(f"   Con testo completo: {database['metadata']['con_testo_completo']}")
    print(f"   Temi indicizzati: {len(database['indice_tematico'])}")
    print(f"\n   Per tag:")
    for tag, count in list(database['metadata']['per_tag'].items())[:10]:
        print(f"     {tag}: {count}")


if __name__ == '__main__':
    main()
